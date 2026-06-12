"""Shared Excel style utilities — unified 威发 quotation format.

Standard layout (12 columns A-L):
  Row 1 (merged A-L): Info row — company/client/contact/date
  Row 2 (merged A-L): Title row — yellow background
  Row 3:              Header row — bold, center
  Row 4-N:            Data rows
  Row N+1 (merged):   Total row — Chinese uppercase + numeric
  Row N+2 (merged):   Note row
  Row N+3 (merged):   Footer row
"""

from datetime import date
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils.units import pixels_to_EMU

# ── Fonts ──────────────────────────────────────────────
FONT_INFO = Font(name="微软雅黑", size=9, color="666666")
FONT_TITLE = Font(name="微软雅黑", size=10, bold=True)
FONT_HEADER = Font(name="微软雅黑", size=10, bold=True)
FONT_DATA = Font(name="微软雅黑", size=11, bold=True)
FONT_TOTAL = Font(name="微软雅黑", size=10, bold=True)
FONT_NOTE = Font(name="微软雅黑", size=10)
FONT_FOOTER = Font(name="微软雅黑", size=9, color="888888")

# ── Fills ──────────────────────────────────────────────
FILL_TITLE = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_NONE = PatternFill(start_color="00000000", end_color="00000000", fill_type=None)

# ── Borders ────────────────────────────────────────────
SIDE_THIN = Side(style="thin")
BORDER_THIN = Border(left=SIDE_THIN, right=SIDE_THIN, top=SIDE_THIN, bottom=SIDE_THIN)

# ── Alignments ─────────────────────────────────────────
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

# ── Column widths (12 columns A-L) ─────────────────────
COLUMN_WIDTHS = {
    "A": 9.66, "B": 27.16, "C": 18.83, "D": 20.16,
    "E": 60.16, "F": 13.33, "G": 7.5, "H": 11.33,
    "I": 6.5, "J": 12.16, "K": 18.16, "L": 16.0,
}

MAX_COL = 12

# ── Number formats ─────────────────────────────────────
NUM_FMT_CURRENCY = '¥#,##0'     # 人民币，0位小数
NUM_FMT_NUMBER = '0'            # 整数
NUM_FMT_PERCENT = '0%'          # 百分比，0位小数


def apply_column_widths(ws):
    """Set standard 12-column widths on worksheet."""
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def _merge_and_style(ws, row, col_start, col_end, value, font, fill, alignment):
    """Write a merged cell row and style it."""
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = BORDER_THIN
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        # Apply border to all merged cells
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).border = BORDER_THIN
            ws.cell(row=row, column=c).fill = fill


def apply_info_row(ws, row, text):
    """Row 1: company/client/contact/date info, gray font, merged A-L."""
    _merge_and_style(ws, row, 1, MAX_COL, text, FONT_INFO, FILL_NONE, ALIGN_LEFT)
    ws.row_dimensions[row].height = 17


def apply_title_row(ws, row, text):
    """Row 2: title, yellow background, merged A-L."""
    _merge_and_style(ws, row, 1, MAX_COL, text, FONT_TITLE, FILL_TITLE, ALIGN_CENTER)
    ws.row_dimensions[row].height = 18


def apply_header_row(ws, row, headers):
    """Row 3: column headers, bold, center, border all sides."""
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_NONE
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
    ws.row_dimensions[row].height = 17


def apply_data_row(ws, row, values, formats=None):
    """Write a data row with standard font, border, center alignment.
    formats: optional dict of {column_index: number_format_string}."""
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = FONT_DATA
        cell.fill = FILL_NONE
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
        if formats and i in formats:
            cell.number_format = formats[i]
    ws.row_dimensions[row].height = 54


def apply_total_row(ws, row, chinese_amount, col_letter="J", data_start_row=4):
    """
    Total row: A-I merged for Chinese uppercase (right align),
    col_letter for SUM formula (center align). Row height 22.
    """
    col_idx = ord(col_letter) - ord("A") + 1
    last_data_row = row - 1
    formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})"
    # A-I: Chinese uppercase
    _merge_and_style(ws, row, 1, 9, chinese_amount, FONT_TOTAL, FILL_NONE, ALIGN_RIGHT)
    # SUM formula
    cell = ws.cell(row=row, column=col_idx, value=formula)
    cell.font = FONT_TOTAL
    cell.fill = FILL_NONE
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER_THIN
    cell.number_format = NUM_FMT_CURRENCY
    # Remaining cols to L: border only
    for c in range(col_idx + 1, 13):
        bc = ws.cell(row=row, column=c)
        bc.border = BORDER_THIN
        bc.fill = FILL_NONE
    ws.row_dimensions[row].height = 22


def apply_note_row(ws, row, text):
    """Note row, merged A-L, left align, regular font."""
    _merge_and_style(ws, row, 1, MAX_COL, text, FONT_NOTE, FILL_NONE, ALIGN_LEFT)
    ws.row_dimensions[row].height = 18


def apply_footer_row(ws, row, text):
    """Footer row, merged A-L, gray font, no bottom border."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_FOOTER
    cell.fill = FILL_NONE
    cell.alignment = ALIGN_LEFT
    # Top-only thin border (no bottom)
    cell.border = Border(top=SIDE_THIN)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
    ws.row_dimensions[row].height = 30


# ── Image embedding ───────────────────────────────────

def embed_image(ws, row, col, image_source: str, upload_dir: str = None):
    """
    Embed a product image into a cell, proportionally resized and centered.
    image_source: URL path (e.g. /product-db/api/uploads/xxx.png) or file path.
    upload_dir: local directory where uploaded files are stored.
    Returns True if image was embedded, False otherwise.
    """
    if not image_source:
        return False

    img_bytes = _resolve_image(image_source, upload_dir)
    if not img_bytes:
        return False

    try:
        from PIL import Image as PILImage
        pil_img = PILImage.open(img_bytes)
        orig_w, orig_h = pil_img.size
        img_bytes.seek(0)
    except Exception as e:
        import logging; logging.getLogger("uvicorn").warning(f"embed_image PIL error: {e}")
        return False

    # Cell dimensions (approximate):
    # Row height 54pt ≈ 72px, column width 16 ≈ 112px
    max_w = 100  # pixels, leave some margin
    max_h = 60   # pixels

    scale = min(max_w / orig_w, max_h / orig_h, 1.0)
    new_w = int(orig_w * scale)
    new_h = int(orig_h * scale)

    col_letter = chr(ord('A') + col - 1)
    # Center image within cell by offsetting from top-left of cell
    # col width ≈ 16*7=112px, row height=54pt≈72px
    cell_w = 112
    cell_h = 72
    offset_x = pixels_to_EMU((cell_w - new_w) // 2)
    offset_y = pixels_to_EMU((cell_h - new_h) // 2)

    xl_img = XlImage(img_bytes)
    xl_img.width = new_w
    xl_img.height = new_h
    xl_img.anchor = f"{col_letter}{row}"

    # Position image from top-left of cell with centering offsets
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D

    marker = AnchorMarker(col=col - 1, colOff=int(offset_x), row=row - 1, rowOff=int(offset_y))
    size = XDRPositiveSize2D(cx=pixels_to_EMU(new_w), cy=pixels_to_EMU(new_h))
    xl_img.anchor = OneCellAnchor(_from=marker, ext=size)

    ws.add_image(xl_img)
    return True


def _resolve_image(source: str, upload_dir: str = None):
    """Resolve image source to BytesIO. Handles relative paths and local files."""
    import os, requests

    # Already bytes/stream
    if isinstance(source, BytesIO):
        return source

    # Relative upload path: /product-db/api/uploads/xxx.png
    if source.startswith("/product-db/api/uploads/") and upload_dir:
        fname = source.rsplit("/", 1)[-1]
        fpath = os.path.join(upload_dir, fname)
        if os.path.isfile(fpath):
            with open(fpath, "rb") as f:
                return BytesIO(f.read())

    # Absolute URL
    if source.startswith("http://") or source.startswith("https://"):
        try:
            r = requests.get(source, timeout=5)
            if r.status_code == 200:
                return BytesIO(r.content)
        except Exception as e:
            import logging; logging.getLogger("uvicorn").warning(f"Failed to download image from {source}: {e}")

    # Local file path
    if os.path.isfile(source):
        with open(source, "rb") as f:
            return BytesIO(f.read())

    return None


# ── Chinese financial uppercase ────────────────────────

_CN_DIGITS = "零壹贰叁肆伍陆柒捌玖"
_CN_RADICES = ["", "拾", "佰", "仟", "万", "拾", "佰", "仟", "亿"]


def num_to_chinese_uppercase(amount):
    """Convert a number to Chinese financial uppercase string. e.g. 30898 → '叁万零捌佰玖拾捌圆整'."""
    if amount is None or amount == 0:
        return "零圆整"
    n = int(round(float(amount)))
    if n < 0:
        return "负" + num_to_chinese_uppercase(-n)

    s = str(n)
    length = len(s)
    result = ""
    zero_flag = False

    for i, ch in enumerate(s):
        pos = length - i - 1  # position from right
        radix_idx = pos % 8
        digit = int(ch)
        radix = _CN_RADICES[radix_idx]

        if digit == 0:
            zero_flag = True
            # insert 万/亿 when crossing section boundaries
            if radix_idx == 4 and pos >= 4:
                result += "万"
                zero_flag = False
            elif radix_idx == 0 and pos >= 8:
                result += "亿"
                zero_flag = False
        else:
            if zero_flag:
                result += "零"
                zero_flag = False
            result += _CN_DIGITS[digit] + radix

    if not result:
        result = "零"

    return result.rstrip("零") + "圆整"
