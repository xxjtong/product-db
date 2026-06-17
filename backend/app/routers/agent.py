"""Hermes Agent proxy — SSE streaming passthrough to Hermes API server.

Frontend calls this endpoint; backend forwards to Hermes and streams back the
OpenAI-compatible SSE response.  Auth via JWT (same as all other API routes).
"""

from __future__ import annotations

import json
import logging
import uuid

import httpx
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse

from app.auth import get_current_user
from app.config import settings, DB_FILESYSTEM_PATH
from app.database import get_db
from app.services.storage import save_file, UPLOAD_DIR
from app.services.approval_manager import approval_manager

logger = logging.getLogger(__name__)

router = APIRouter()

HERMES_CHAT_URL = f"{settings.HERMES_API_URL.rstrip('/')}/v1/chat/completions"
HERMES_TIMEOUT = httpx.Timeout(300.0, connect=10.0)

_AGENT_PROMPT_DEFAULT = "你是 pdb，产品数据库系统的 AI 助手。"

ALLOWED_UPLOAD_TYPES = {
    "image/png", "image/jpeg", "image/jpg", "image/webp", "image/gif", "image/bmp",
    "text/plain", "text/csv", "application/json", "text/markdown",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # xlsx
    "application/vnd.ms-excel",  # xls
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",  # docx
    "text/xml", "application/xml",
}
MAX_UPLOAD_SIZE = 20 * 1024 * 1024  # 20MB

EXTRACTABLE_TYPES = {
    "text/plain", "text/csv", "application/json", "text/markdown",
    "text/xml", "application/xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}


def _extract_file_text(path, filename: str, mime_type: str) -> str:
    """Extract text content from uploaded file for LLM context."""
    if mime_type in ("text/plain", "text/csv", "application/json", "text/markdown", "text/xml", "application/xml"):
        return path.read_text(errors="replace")[:50000]

    if "excel" in mime_type or "spreadsheet" in mime_type or filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            lines = []
            for name in wb.sheetnames:
                ws = wb[name]
                lines.append(f"=== Sheet: {name} ===")
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(c.strip() for c in cells):
                        lines.append(" | ".join(cells))
                lines.append("")
            return "\n".join(lines)[:50000]
        except Exception as e:
            return f"[无法解析Excel文件: {e}]"

    if "pdf" in mime_type or filename.endswith(".pdf"):
        try:
            import subprocess, tempfile
            result = subprocess.run(["pdftotext", "-layout", str(path), "-"], capture_output=True, text=True, timeout=30)
            return result.stdout[:50000] or "[PDF提取结果为空]"
        except Exception as e:
            return f"[无法解析PDF文件: {e}]"

    if "wordprocessing" in mime_type or filename.endswith(".docx"):
        try:
            from docx import Document
            doc = Document(str(path))
            return "\n".join(p.text for p in doc.paragraphs)[:50000]
        except Exception as e:
            return f"[无法解析DOCX文件: {e}]"

    return ""


def _get_agent_prompt(db):
    """Load agent_prompt from system_settings, fallback to import default."""
    from app.models.system_setting import SystemSetting
    s = db.query(SystemSetting).filter_by(key="agent_prompt").first()
    if s and s.value:
        return s.value
    # Fallback to the module-level default in admin_routes
    try:
        from app.routers.admin_routes import _PROMPT_DEFAULTS
        return _PROMPT_DEFAULTS.get("agent_prompt", _AGENT_PROMPT_DEFAULT)
    except Exception:
        return _AGENT_PROMPT_DEFAULT


@router.get("/agent/config")
async def agent_config(user=Depends(get_current_user)):
    """Return agent configuration including database path and API base URL."""
    api_base = settings.AGENT_API_BASE or f"http://127.0.0.1:{8000 if settings.DEV_MODE else 8002}"
    return {
        "db_path": DB_FILESYSTEM_PATH,
        "api_base": f"{api_base}/product-db/api",
    }


@router.get("/agent/prompt")
async def agent_prompt(user=Depends(get_current_user), db=Depends(get_db)):
    """Return the Hermes agent system prompt (admin-editable, DB-backed)."""
    return {"prompt": _get_agent_prompt(db)}


def _build_auth_header() -> dict:
    """Return Authorization header dict if an API key is configured."""
    key = settings.HERMES_API_KEY.strip()
    if key:
        return {"Authorization": f"Bearer {key}"}
    return {}


@router.post("/agent/upload")
async def agent_upload(
    request: Request,
    user=Depends(get_current_user),
):
    """Upload a file for Agent context. Returns a public URL Hermes can fetch."""
    content_type = request.headers.get("content-type", "")
    if not content_type.startswith("multipart/form-data"):
        raise HTTPException(400, "Expected multipart/form-data")

    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(400, "No file provided")

    import mimetypes
    guessed = mimetypes.guess_type(file.filename or "")[0]
    mime_type = file.content_type or ""
    # Prefer guessed type over generic octet-stream from browser/curl
    if not mime_type or mime_type == "application/octet-stream":
        mime_type = guessed or mime_type or "application/octet-stream"

    if mime_type not in ALLOWED_UPLOAD_TYPES:
        raise HTTPException(400, f"File type not allowed: {mime_type}")

    contents = await file.read()
    if len(contents) > MAX_UPLOAD_SIZE:
        raise HTTPException(400, f"File too large (max {MAX_UPLOAD_SIZE // 1024 // 1024}MB)")

    # Save to uploads dir with UUID filename
    import uuid
    ext = (file.filename or "file").rsplit(".", 1)[-1] if "." in (file.filename or "") else ""
    stored_name = f"{uuid.uuid4().hex}"
    if ext:
        stored_name += f".{ext}"
    stored_path = UPLOAD_DIR / stored_name
    stored_path.write_bytes(contents)

    # Build public URL — use configured API base or derive from request
    from urllib.parse import urlparse
    api_base = settings.AGENT_API_BASE
    if api_base:
        base_host = api_base.rstrip("/")
    else:
        base_host = str(request.base_url).rstrip("/")

    file_url = f"{base_host}/product-db/api/uploads/{stored_name}"

    logger.info("agent_upload: %s → %s (%d bytes)", file.filename, stored_name, len(contents))
    return {
        "url": file_url,
        "filename": file.filename,
        "size": len(contents),
        "type": mime_type,
    }


@router.post("/agent/approval/{task_id}")
async def agent_approval(
    task_id: str,
    request: Request,
    user=Depends(get_current_user),
):
    """Human decision on a pending agent tool call."""
    body = await request.json()
    approved = body.get("approved", False)
    reason = body.get("reason", "")

    ok = approval_manager.decide(task_id, approved, reason)
    if not ok:
        raise HTTPException(404, "Approval task not found")
    return {"ok": True, "task_id": task_id}


@router.get("/agent/approvals")
async def agent_approvals(user=Depends(get_current_user)):
    """List pending approval tasks."""
    pending = approval_manager.get_pending()
    return {
        "tasks": [
            {
                "task_id": t.task_id,
                "tool_name": t.tool_name,
                "tool_label": t.tool_label,
                "summary": t.summary,
                "details": t.details,
                "tool_input": t.tool_input,
            }
            for t in pending
        ]
    }


@router.post("/agent/test-approval")
async def agent_test_approval(user=Depends(get_current_user)):
    """Create a test approval task for UI testing."""
    task = approval_manager.create(
        tool_name="create_quotation",
        tool_label="创建报价单",
        tool_input={"solution_id": 26, "total_price": 636301.50, "customer": "岭南大学"},
        summary="为方案「岭南大学新科学楼 IoT」创建报价单，总价 ¥636,301.50",
        details={"message": "已根据 IoT QTY 需求匹配 11 款产品"},
    )
    return {"task_id": task.task_id}


# Write-operation tool names that require human approval
_WRITE_TOOLS = {
    "create_quotation": "创建报价单",
    "save_bom": "保存BOM",
    "update_solution": "修改方案",
    "delete_product": "删除产品",
}


def _needs_approval(tool_name: str) -> str:
    """Return human-readable label if tool needs approval, empty string otherwise."""
    return _WRITE_TOOLS.get(tool_name, "")


@router.post("/agent/chat")
async def agent_chat(
    request: Request,
    user=Depends(get_current_user),
):
    """Proxy a chat request to Hermes, intercepting write-op tool calls for approval.

    Body (JSON):
        messages: list of {"role": "...", "content": "..."}
        stream: true/false (default true for SSE streaming)
        model: optional model name override

    Returns: SSE text/event-stream from Hermes /v1/chat/completions,
             with approval_required events injected for write-op tool calls.
    """
    body = await request.json()

    messages = body.get("messages")
    if not messages or not isinstance(messages, list):
        raise HTTPException(status_code=400, detail="Missing or invalid 'messages' field")

    stream = body.get("stream", True)
    model = body.get("model", "hermes-agent")

    payload = {
        "model": model,
        "messages": messages,
        "stream": False,
    }

    headers = {
        "Content-Type": "application/json",
        **_build_auth_header(),
    }

    logger.info("agent_chat: proxying %d messages to %s", len(messages), HERMES_CHAT_URL)

    async def _stream():
        try:
            async with httpx.AsyncClient(timeout=HERMES_TIMEOUT) as client:
                # First: get Hermes response (non-streaming to detect tool_calls)
                resp = await client.post(HERMES_CHAT_URL, json=payload, headers=headers)

                if resp.status_code != 200:
                    body_text = await resp.aread()
                    logger.warning("agent_chat: Hermes returned %s: %s", resp.status_code, body_text[:500])
                    yield f"data: {json.dumps({'error': f'Hermes returned {resp.status_code}'})}\n\n"
                    yield "data: [DONE]\n\n"
                    return

                data = resp.json()
                choice = (data.get("choices") or [{}])[0]
                message = choice.get("message", {})
                tool_calls = message.get("tool_calls") or []

                # Test trigger: inject approval for "测试审批" in user message
                if not tool_calls:
                    last_user_msg = ""
                    for m in reversed(messages):
                        if m.get("role") == "user":
                            last_user_msg = m.get("content", "")
                            break
                    if isinstance(last_user_msg, list):
                        last_user_msg = " ".join(p.get("text", "") for p in last_user_msg if p.get("type") == "text")
                    if "测试审批" in last_user_msg:
                        tool_calls = [{
                            "id": "test_" + uuid.uuid4().hex[:8],
                            "type": "function",
                            "function": {
                                "name": "create_quotation",
                                "arguments": json.dumps({"solution_id": 26, "total_price": 636301.50, "customer": "岭南大学"}),
                            },
                        }]

                # Check for write-op tool calls that need human approval
                approval_tasks = []
                for tc in tool_calls:
                    func = tc.get("function", {})
                    tool_name = func.get("name", "")
                    label = _needs_approval(tool_name)
                    if label:
                        try:
                            tool_input = json.loads(func.get("arguments", "{}"))
                        except (json.JSONDecodeError, TypeError):
                            tool_input = {}
                        summary = f"{label}: {json.dumps(tool_input, ensure_ascii=False)[:200]}"
                        task = approval_manager.create(
                            tool_name=tool_name,
                            tool_label=label,
                            tool_input=tool_input,
                            summary=summary,
                            details={
                                "message": message.get("content", "")[:500],
                                "tool_call_id": tc.get("id", ""),
                            },
                        )
                        approval_tasks.append(task)

                if approval_tasks:
                    # Emit approval_required events before waiting
                    for task in approval_tasks:
                        approval_event = {
                            "type": "approval_required",
                            "task_id": task.task_id,
                            "tool_name": task.tool_name,
                            "tool_label": task.tool_label,
                            "summary": task.summary,
                            "details": task.details,
                            "tool_input": task.tool_input,
                        }
                        yield f"data: {json.dumps(approval_event, ensure_ascii=False)}\n\n"
                        logger.info("agent_chat: emitted approval_required for %s", task.task_id)

                    # Wait for all approvals, collect decisions
                    decisions = []
                    for task in approval_tasks:
                        decision = await approval_manager.wait_for_decision(task.task_id)
                        decisions.append(decision)
                        result_event = {
                            "type": "approval_result",
                            "task_id": task.task_id,
                            "approved": decision.get("approved", False),
                            "reason": decision.get("reason", ""),
                        }
                        yield f"data: {json.dumps(result_event, ensure_ascii=False)}\n\n"

                    # Append tool result messages based on decisions
                    any_approved = False
                    for task, decision in zip(approval_tasks, decisions):
                        if decision.get("approved"):
                            any_approved = True
                            messages.append({
                                "role": "tool",
                                "tool_call_id": task.details.get("tool_call_id", ""),
                                "content": json.dumps({"status": "approved", "message": "用户已授权此操作"}),
                            })
                        else:
                            messages.append({
                                "role": "tool",
                                "tool_call_id": task.details.get("tool_call_id", ""),
                                "content": json.dumps({"error": f"操作被用户拒绝: {decision.get('reason', '未说明原因')}"}),
                            })

                    if any_approved:
                        # Re-call Hermes with tool results
                        payload2 = {"model": model, "messages": messages, "stream": stream}
                        async with client.stream("POST", HERMES_CHAT_URL, json=payload2, headers=headers) as resp2:
                            if resp2.status_code == 200:
                                async for chunk in resp2.aiter_bytes():
                                    yield chunk
                            else:
                                body2 = await resp2.aread()
                                yield f"data: {json.dumps({'error': f'Hermes returned {resp2.status_code} on retry'})}\n\n"
                                yield "data: [DONE]\n\n"
                    else:
                        # All rejected — return the original response with rejection note
                        data["choices"][0]["message"]["content"] = (message.get("content", "") + "\n\n⚠️ 操作已被用户拒绝。").strip()
                        yield json.dumps(data)
                elif stream:
                    # No approval needed — re-fetch with streaming for real-time display
                    payload_s = {"model": model, "messages": messages, "stream": True}
                    async with client.stream("POST", HERMES_CHAT_URL, json=payload_s, headers=headers) as resp_s:
                        if resp_s.status_code == 200:
                            async for chunk in resp_s.aiter_bytes():
                                yield chunk
                        else:
                            # Fallback: return the non-streaming response as SSE
                            yield f"data: {json.dumps(data)}\n\n"
                            yield "data: [DONE]\n\n"
                else:
                    # Non-streaming, no approval
                    yield json.dumps(data)
                    yield "\n"
        except httpx.ConnectError:
            logger.exception("agent_chat: cannot connect to Hermes at %s", HERMES_CHAT_URL)
            yield f"data: {json.dumps({'error': 'Cannot connect to Hermes agent server'})}\n\n"
            yield "data: [DONE]\n\n"
        except Exception:
            logger.exception("agent_chat: stream error")
            yield f"data: {json.dumps({'error': 'Agent stream interrupted'})}\n\n"
            yield "data: [DONE]\n\n"

    return StreamingResponse(
        _stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )
