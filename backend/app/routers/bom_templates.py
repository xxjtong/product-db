"""BOM Template CRUD + Solution BOM Snapshot management + xlsx export."""
from __future__ import annotations
import json
import re
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.utils.helpers import get_or_404, apply_partial_update, format_description_with_specs
from app.models.bom_template import BOMTemplate, SolutionBOMSnapshot
from app.models.solution import Solution, SolutionItem
from app.models.product import Product
from app.auth import get_current_user, check_ownership
from app.models.user import User
from app.schemas.bom_template import BOMTemplateCreate, BOMTemplateUpdate, SaveAsTemplateRequest
from app.schemas.solution import BOMSnapshotSave
from datetime import datetime, timezone
from io import BytesIO

router = APIRouter()


# --- BOM Templates ---

@router.get("/bom-templates")
def list_templates(db: Session = Depends(get_db), user=Depends(get_current_user)):
    templates = db.query(BOMTemplate).order_by(BOMTemplate.name).all()
    return {"templates": [t.to_dict() for t in templates]}


@router.post("/bom-templates", status_code=201)
def create_template(data: BOMTemplateCreate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    t = BOMTemplate(
        name=data.name,
        description=data.description,
        sheet_name=data.sheet_name,
        snapshot=data.snapshot,
        is_default=data.is_default,
        created_by=user.id,
    )
    db.add(t)
    db.commit()
    db.refresh(t)
    return {"template": t.to_dict()}


@router.get("/bom-templates/{template_id}")
def get_template(template_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    t = get_or_404(db, BOMTemplate, template_id, "Template not found")
    return {"template": t.to_dict()}


@router.put("/bom-templates/{template_id}")
def update_template(template_id: int, data: BOMTemplateUpdate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    t = get_or_404(db, BOMTemplate, template_id, "Template not found")
    check_ownership(t, user)
    apply_partial_update(t, data, ["name", "description", "sheet_name", "snapshot", "is_default"])
    t.updated_at = datetime.now(timezone.utc)
    db.commit()
    return {"template": t.to_dict()}


@router.delete("/bom-templates/{template_id}")
def delete_template(template_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    t = get_or_404(db, BOMTemplate, template_id, "Template not found")
    check_ownership(t, user, strict=True)
    db.delete(t)
    db.commit()
    return {"ok": True}


@router.post("/bom-templates/{template_id}/duplicate", status_code=201)
def duplicate_template(template_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    t = get_or_404(db, BOMTemplate, template_id, "Template not found")
    new_t = BOMTemplate(
        name=f"{t.name} (副本)",
        description=t.description,
        sheet_name=t.sheet_name,
        snapshot=t.snapshot,
        is_default=False,
        created_by=user.id,
    )
    db.add(new_t)
    db.commit()
    db.refresh(new_t)
    return {"template": new_t.to_dict()}


# --- Solution BOM Snapshot ---

@router.get("/solutions/{solution_id}/bom-snapshot")
def get_bom_snapshot(solution_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    sol = get_or_404(db, Solution, solution_id, "Solution not found")
    check_ownership(sol, user, strict=False)

    existing = db.query(SolutionBOMSnapshot).filter_by(solution_id=solution_id).first()
    if existing:
        return {"bom_snapshot": existing.to_dict()}

    # Generate from template + solution items
    template_id = None
    default_template = db.query(BOMTemplate).filter_by(is_default=True).first()
    if not default_template:
        default_template = db.query(BOMTemplate).first()
    template_id = default_template.id if default_template else None

    snapshot_data = _generate_snapshot(sol, default_template, db)
    new_snapshot = SolutionBOMSnapshot(
        solution_id=solution_id,
        template_id=template_id,
        snapshot=snapshot_data,
    )
    db.add(new_snapshot)
    db.commit()
    db.refresh(new_snapshot)
    return {"bom_snapshot": new_snapshot.to_dict()}


@router.put("/solutions/{solution_id}/bom-snapshot")
def save_bom_snapshot(solution_id: int, data: BOMSnapshotSave, db: Session = Depends(get_db), user=Depends(get_current_user)):
    sol = get_or_404(db, Solution, solution_id, "Solution not found")
    check_ownership(sol, user, strict=True)

    existing = db.query(SolutionBOMSnapshot).filter_by(solution_id=solution_id).first()
    if existing:
        existing.snapshot = data.snapshot if data.snapshot else existing.snapshot
        existing.updated_at = datetime.now(timezone.utc)
    else:
        existing = SolutionBOMSnapshot(
            solution_id=solution_id,
            template_id=None,
            snapshot=data.snapshot,
        )
        db.add(existing)
    db.commit()
    db.refresh(existing)

    # Sync snapshot cell data back to solution_items
    _sync_snapshot_to_items(solution_id, data.snapshot, db)

    return {"bom_snapshot": existing.to_dict()}


def _sync_snapshot_to_items(solution_id: int, snapshot: dict, db: Session):
    """Parse BOM snapshot cells and update solution_items with edited values.

    Sheet layout: data rows start at row 3 (snapshot) or row 5 (basic bom).
    Columns: A=# B=名称 C=型号/SKU D=功能描述 E=数量 F=单价 G=折扣% H=小计 I=备注 J=成本
    We match by SKU in column C to find the corresponding solution_item.
    """
    cells = snapshot.get("cells", {})
    if not cells:
        return

    # Build ref → (row, col) lookup, grouped by row
    rows = {}
    for ref, cell in cells.items():
        if not isinstance(cell, dict):
            continue
        m = re.match(r"^([A-Z]+)(\d+)$", ref)
        if not m:
            continue
        col_letter = m.group(1)
        row_num = int(m.group(2))
        if row_num not in rows:
            rows[row_num] = {}
        rows[row_num][col_letter] = cell

    # Extract SKU → values mapping from data rows
    sku_updates = {}
    for row_num in sorted(rows.keys()):
        rd = rows[row_num]
        sku = rd.get("C", {}).get("v")
        if not sku or not isinstance(sku, str) or not sku.strip():
            continue
        sku = sku.strip()
        qty = float(rd.get("E", {}).get("v", 0) or 0)
        price = float(rd.get("F", {}).get("v", 0) or 0)
        discount = float(rd.get("G", {}).get("v", 100) or 100)
        remark = str(rd.get("I", {}).get("v", "") or "")
        sku_updates[sku] = {"quantity": qty, "unit_price": price, "discount_rate": discount, "remark": remark}

    if not sku_updates:
        return

    # Batch-load products for all solution items to map SKU
    items = db.query(SolutionItem).filter_by(solution_id=solution_id).all()
    pids = [it.product_id for it in items if it.product_id]
    products = {p.id: p for p in db.query(Product).filter(Product.id.in_(pids)).all()} if pids else {}

    updated = 0
    for item in items:
        p = products.get(item.product_id)
        if not p or not p.sku:
            continue
        u = sku_updates.get(p.sku)
        if not u:
            continue
        item.quantity = u["quantity"]
        item.unit_price = u["unit_price"]
        item.discount_rate = u["discount_rate"]
        if u["remark"]:
            item.remark = u["remark"]
        updated += 1

    if updated:
        db.commit()


@router.post("/solutions/{solution_id}/bom-snapshot/save-as-template", status_code=201)
def save_as_template(solution_id: int, data: SaveAsTemplateRequest, db: Session = Depends(get_db), user=Depends(get_current_user)):
    sol = get_or_404(db, Solution, solution_id, "Solution not found")
    check_ownership(sol, user, strict=False)
    snap = db.query(SolutionBOMSnapshot).filter_by(solution_id=solution_id).first()
    if not snap:
        raise HTTPException(404, "No BOM snapshot found for this solution")

    t = BOMTemplate(
        name=data.name or f"Template from Solution #{solution_id}",
        description=data.description,
        sheet_name=data.sheet_name,
        snapshot=snap.snapshot,
        is_default=False,
        created_by=user.id,
    )
    db.add(t)
    db.commit()
    db.refresh(t)
    return {"template": t.to_dict()}


@router.get("/solutions/{solution_id}/bom-snapshot/export-xlsx")
def export_bom_xlsx(solution_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    sol = get_or_404(db, Solution, solution_id, "Solution not found")
    check_ownership(sol, user, strict=False)

    snap = db.query(SolutionBOMSnapshot).filter_by(solution_id=solution_id).first()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    if snap and snap.snapshot and snap.snapshot.get("cells"):
        _write_snapshot_to_xlsx(ws, snap.snapshot)
    else:
        _write_basic_bom(ws, sol, solution_id, db)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"bom_solution_{solution_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


def _cell_ref_to_rc(ref: str):
    """Convert 'A1' to (row, col) tuple (1-indexed)."""
    m = re.match(r'^([A-Z]+)(\d+)$', ref)
    if not m:
        return 1, 1
    col = 0
    for ch in m.group(1):
        col = col * 26 + (ord(ch) - 64)
    return int(m.group(2)), col


def _inline_style_to_openpyxl(s: dict) -> dict:
    """Convert snapshot inline style to openpyxl kwargs."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    result = {}
    font_kwargs = {}

    if s.get("fs"):
        font_kwargs["size"] = s["fs"]
    if s.get("bl"):
        font_kwargs["bold"] = True
    if s.get("it"):
        font_kwargs["italic"] = True
    if s.get("ul"):
        font_kwargs["underline"] = "single"
    if s.get("ff"):
        font_kwargs["name"] = s["ff"]
    if s.get("cl"):
        cl = s["cl"]
        if isinstance(cl, dict):
            cl = cl.get("rgb", "#000000")
        font_kwargs["color"] = cl.lstrip("#")

    if font_kwargs:
        result["font"] = Font(**font_kwargs)

    if s.get("bg"):
        result["fill"] = PatternFill(start_color=s["bg"].lstrip("#"),
                                      end_color=s["bg"].lstrip("#"),
                                      fill_type="solid")

    align_kwargs = {}
    if s.get("ht") is not None:
        align_map = {0: "left", 1: "left", 2: "center", 3: "right", "left": "left", "center": "center", "right": "right"}
        align_kwargs["horizontal"] = align_map.get(s["ht"], "left")
    if s.get("vt") is not None:
        v_map = {0: "top", 1: "top", 2: "center", 3: "bottom", "top": "top", "middle": "center", "bottom": "bottom"}
        align_kwargs["vertical"] = v_map.get(s["vt"], "top")
    if s.get("tb"):
        align_kwargs["wrap_text"] = True
    if align_kwargs:
        result["alignment"] = Alignment(**align_kwargs)

    return result


def _write_snapshot_to_xlsx(ws, snapshot: dict):
    """Write snapshot data (cells, styles, merges, colWidths, rowHeights) to openpyxl worksheet."""
    cells = snapshot.get("cells", {})

    # Write cell values, formulas, and styles
    for ref, cell in cells.items():
        row, col = _cell_ref_to_rc(ref)
        openpyxl_cell = ws.cell(row=row, column=col)

        v = cell.get("v")
        if v is not None and v != "":
            openpyxl_cell.value = v

        if cell.get("f"):
            openpyxl_cell.value = "=" + cell["f"]

        if cell.get("s"):
            style_kwargs = _inline_style_to_openpyxl(cell["s"])
            for attr, value in style_kwargs.items():
                setattr(openpyxl_cell, attr, value)

    # Apply column widths (keys are cell refs like "A1" from Univer save)
    col_widths = snapshot.get("colWidths", {})
    for col_key, width in col_widths.items():
        col_letter = ''.join(c for c in col_key if c.isalpha())
        ws.column_dimensions[col_letter].width = max(float(width) * 0.14, 3)

    # Apply row heights
    row_heights = snapshot.get("rowHeights", {})
    for row_key, height in row_heights.items():
        ws.row_dimensions[int(row_key)].height = float(height) * 0.75

    # Apply merges
    merges = snapshot.get("merges", [])
    for ref in merges:
        parts = ref.split(":")
        start = _cell_ref_to_rc(parts[0])
        end = _cell_ref_to_rc(parts[1]) if len(parts) > 1 else start
        ws.merge_cells(
            start_row=start[0], start_column=start[1],
            end_row=end[0], end_column=end[1]
        )


def _write_basic_bom(ws, sol, solution_id: int, db: Session):
    """Fallback: generate a BOM sheet from solution_items with unified style."""
    from app.utils.excel_style import (
        apply_info_row, apply_title_row, apply_header_row, apply_data_row,
        apply_total_row, apply_note_row, apply_footer_row, apply_column_widths,
        num_to_chinese_uppercase, NUM_FMT_CURRENCY, NUM_FMT_NUMBER, NUM_FMT_PERCENT,
        embed_image,
    )
    from datetime import date

    ws.title = "BOM清单"
    apply_column_widths(ws)

    # Row 1: info
    today = date.today().isoformat()
    info_text = f"客户：{sol.client_name or ''}  |  项目：{sol.project_name or ''}  |  日期：{today}"
    apply_info_row(ws, 1, info_text)

    # Row 2: title
    apply_title_row(ws, 2, f"BOM清单 — {sol.name}")

    # Row 3: headers
    headers = ["序号", "名称", "规格型号", "型号", "功能描述", "单价", "数量", "合计", "折扣率", "成交价", "备注", "图片"]
    apply_header_row(ws, 3, headers)
    # Cost header (M): plain text, no style — safe to delete column
    ws.cell(row=3, column=13).value = "成本"

    # Data rows
    items = db.query(SolutionItem).filter_by(solution_id=solution_id)\
        .order_by(SolutionItem.sort_order).all()
    pids = {it.product_id for it in items if it.product_id}
    product_map = {p.id: p for p in db.query(Product).filter(Product.id.in_(pids)).all()} if pids else {}
    for idx, item in enumerate(items, 1):
        p = product_map.get(item.product_id)
        qty = float(item.quantity or 0)
        price = float(item.unit_price or 0)
        discount = float(item.discount_rate or 100)
        row = 3 + idx
        formats = {6: NUM_FMT_CURRENCY, 7: NUM_FMT_NUMBER, 8: NUM_FMT_CURRENCY,
                   9: NUM_FMT_PERCENT, 10: NUM_FMT_CURRENCY}
        apply_data_row(ws, row, [
            idx,
            p.name if p else "",
            (p.model or "") if p else "",
            (p.model or "") if p else "",
            format_description_with_specs(p.description or "", p.specs or {}) if p else "",
            price,
            qty,
            price * qty,  # H placeholder
            discount / 100,  # I: 折扣率(小数)
            price * qty * (discount / 100),  # J placeholder
            item.remark or "",
            p.image_url or "" if p else "",
        ], formats)
        # Cost column (M): plain value, no style — safe to delete column
        ws.cell(row=row, column=13).value = float(p.cost_price or 0) if p else 0
        # Replace H and J with formulas
        ws.cell(row=row, column=8).value = f"=F{row}*G{row}"       # H: 合计
        ws.cell(row=row, column=10).value = f"=H{row}*I{row}"       # J: 成交价

        # Embed product image in column L
        if p and p.image_url:
            import os
            from app.services.storage import UPLOAD_DIR
            if not embed_image(ws, row, 12, p.image_url, str(UPLOAD_DIR)):
                import logging; logging.getLogger("uvicorn").warning(f"Failed to embed image for product {p.id} at row {row}")

    # Total row
    total_row = 3 + len(items) + 1
    total = float(sol.total_price or 0)
    apply_total_row(ws, total_row, f"合计（大写）：{num_to_chinese_uppercase(total)}", col_letter="J")

    # Note row
    apply_note_row(ws, total_row + 1, f"BOM清单 — {sol.name}  |  方案编号：{solution_id}")

    # Footer
    apply_footer_row(ws, total_row + 2, "产品数据库 — BOM清单导出")


def _generate_snapshot(sol: Solution, template, db: Session) -> dict:
    """Generate a BOM snapshot by merging template structure with solution items."""
    if template and template.snapshot:
        snapshot = dict(template.snapshot)
    else:
        snapshot = {"cells": {}, "colWidths": {}, "rowHeights": {}}

    items = db.query(SolutionItem).filter_by(solution_id=sol.id)\
        .order_by(SolutionItem.sort_order).all()
    pids = {it.product_id for it in items if it.product_id}
    product_map = {p.id: p for p in db.query(Product).filter(Product.id.in_(pids)).all()} if pids else {}

    cells = snapshot.get("cells", {})
    # Overlay headers matching BOM editor table layout
    cells["A1"] = {"v": "#"}
    cells["B1"] = {"v": "产品名称"}
    cells["C1"] = {"v": "型号/SKU"}
    cells["D1"] = {"v": "功能描述"}
    cells["E1"] = {"v": "数量"}
    cells["F1"] = {"v": "单价"}
    cells["G1"] = {"v": "折扣%"}
    cells["H1"] = {"v": "小计"}
    cells["I1"] = {"v": "备注"}
    cells["J1"] = {"v": "成本"}
    start_row = 3
    for idx, item in enumerate(items):
        row = start_row + idx
        p = product_map.get(item.product_id)
        qty = float(item.quantity or 0)
        price = float(item.unit_price or 0)
        discount = float(item.discount_rate or 100)
        cells[f"A{row}"] = {"v": idx + 1}
        cells[f"B{row}"] = {"v": p.name if p else ""}
        cells[f"C{row}"] = {"v": p.sku if p else ""}
        cells[f"D{row}"] = {"v": (p.description or "")[:200] if p else ""}
        cells[f"E{row}"] = {"v": qty}
        cells[f"F{row}"] = {"v": price}
        cells[f"G{row}"] = {"v": discount}
        cells[f"H{row}"] = {"v": qty * price * (discount / 100)}
        cells[f"I{row}"] = {"v": item.remark or ""}
        cells[f"J{row}"] = {"v": float(p.cost_price or 0) if p else 0}

    # Total row
    total_row = start_row + len(items) + 1
    cells[f"H{total_row}"] = {"v": float(sol.total_price or 0)}

    snapshot["cells"] = cells
    return snapshot
