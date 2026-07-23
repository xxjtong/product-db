"""Quotation CRUD + export."""
from __future__ import annotations
import json
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, selectinload
from sqlalchemy import select
from app.database import get_db
from app.utils.helpers import get_or_404, apply_partial_update, format_description_with_specs
from app.models.quotation import Quotation, QuotationItem
from app.models.product import Product
from app.models.solution import Solution, SolutionItem
from app.auth import get_current_user, filter_by_ownership, check_ownership
from app.models.user import User
from app.utils.escape import escape_like
from app.schemas.quotation import QuotationCreate, QuotationUpdate, QuotationItemCreate, QuotationItemUpdate
from app.schemas.solution import BatchDeleteRequest
from datetime import datetime, timezone
from io import BytesIO

router = APIRouter()


def _generate_quote_number(db: Session) -> str:
    """Generate quote number: QT-YYYYMMDD-NNN."""
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    prefix = f"QT-{today}-"
    last = db.query(Quotation).filter(Quotation.quote_number.like(f"{prefix}%"))\
        .order_by(Quotation.id.desc()).with_for_update().first()
    seq = 1
    if last and last.quote_number:
        try:
            seq = int(last.quote_number.split("-")[-1]) + 1
        except (ValueError, IndexError):
            seq = 1
    return f"{prefix}{seq:03d}"


def _recalc_total(qt: Quotation, db: Session):
    items = db.query(QuotationItem).filter_by(quotation_id=qt.id).all()
    total = 0
    for item in items:
        qty = float(item.quantity or 0)
        price = float(item.unit_price or 0)
        rate = float(item.discount_rate or 100)
        amount = qty * price * (rate / 100)
        item.amount = amount
        total += amount
    qt.total_amount = total


@router.get("/quotations")
def list_quotations(
    status: Optional[str] = None,
    search: str = "",
    page: int = 1,
    per_page: int = 20,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    q = db.query(Quotation).options(
        selectinload(Quotation.items),
    )
    q = filter_by_ownership(q, Quotation, user, strict=True)
    if status:
        q = q.filter(Quotation.status == status)
    if search:
        q = q.filter(
            Quotation.quote_number.ilike(f"%{escape_like(search)}%")
            | Quotation.title.ilike(f"%{escape_like(search)}%")
            | Quotation.client_name.ilike(f"%{escape_like(search)}%")
        )
    from app.utils.helpers import paginate
    quotations, total = paginate(q.order_by(Quotation.updated_at.desc()), page, per_page)
    return {
        "quotations": [qt.to_dict() for qt in quotations],
        "total": total,
        "page": page,
        "per_page": per_page,
    }


@router.post("/quotations", status_code=201)
def create_quotation(data: QuotationCreate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    qt = Quotation(
        solution_id=data.solution_id,
        quote_number=_generate_quote_number(db),
        title=data.title,
        client_name=data.client_name,
        client_contact=data.client_contact,
        valid_days=data.valid_days,
        tax_rate=data.tax_rate,
        status=data.status,
        notes=data.notes,
        created_by=user.id,
    )

    # Populate from solution if provided
    solution_id = data.solution_id
    if solution_id:
        sol = db.get(Solution, solution_id)
        if sol:
            if not qt.title:
                qt.title = sol.name
            if not qt.client_name:
                qt.client_name = sol.client_name
            sol_items = db.query(SolutionItem).options(
                selectinload(SolutionItem.product),
            ).filter_by(solution_id=solution_id).all()
            # Preload products with all relationships for full snapshots
            product_ids = [si.product_id for si in sol_items]
            products_map = {}
            if product_ids:
                prods = db.query(Product).options(
                    selectinload(Product.category),
                    selectinload(Product.manufacturer),
                    selectinload(Product.supplier),
                    selectinload(Product.comm_methods),
                    selectinload(Product.comm_protocols),
                    selectinload(Product.power_supplies),
                    selectinload(Product.hardware_interfaces),
                    selectinload(Product.sensor_capabilities),
                    selectinload(Product.images),
                ).filter(Product.id.in_(product_ids)).all()
                products_map = {p.id: p for p in prods}
            for idx, si in enumerate(sol_items):
                prod = products_map.get(si.product_id)
                snapshot = prod.to_dict() if prod else {}
                qi = QuotationItem(
                    quotation=qt,
                    solution_item_id=si.id,
                    product_id=si.product_id,
                    product_snapshot=snapshot,
                    quantity=si.quantity,
                    unit_price=si.unit_price or (float(prod.base_price) if prod and prod.base_price else 0),
                    amount=float(si.quantity or 0) * float(si.unit_price or 0),
                    discount_rate=si.discount_rate,
                    remark=si.remark,
                    sort_order=idx + 1,
                )
                db.add(qi)

    db.add(qt)
    db.commit()
    db.refresh(qt)
    _recalc_total(qt, db)
    db.commit()
    db.refresh(qt)
    return {"quotation": qt.to_dict()}


@router.post("/quotations/batch-delete")
def batch_delete_quotations(data: BatchDeleteRequest, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not data.ids:
        raise HTTPException(400, "ids is required")
    if user.role != "admin":
        owned = db.query(Quotation.id).filter(
            Quotation.id.in_(data.ids),
            Quotation.created_by == user.id,
        ).all()
        owned_ids = {r[0] for r in owned}
        forbidden = [i for i in data.ids if i not in owned_ids]
        if forbidden:
            raise HTTPException(403, f"Access denied for quotations: {forbidden}")
    deleted = db.query(Quotation).filter(Quotation.id.in_(data.ids)).delete(synchronize_session="fetch")
    db.commit()
    return {"ok": True, "deleted": deleted}


@router.get("/quotations/{quotation_id}")
def get_quotation(quotation_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    qt = db.scalar(select(Quotation).options(
        selectinload(Quotation.items),
    ).where(Quotation.id == quotation_id))
    if not qt:
        raise HTTPException(404, "Quotation not found")
    check_ownership(qt, user, strict=True)
    return {"quotation": qt.to_dict()}


@router.put("/quotations/{quotation_id}")
def update_quotation(quotation_id: int, data: QuotationUpdate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    qt = get_or_404(db, Quotation, quotation_id, "Quotation not found")
    check_ownership(qt, user, strict=True)
    apply_partial_update(qt, data, ["title", "client_name", "client_contact", "valid_days", "tax_rate", "status", "notes"])
    qt.updated_at = datetime.now(timezone.utc)
    db.commit()
    qt = db.scalar(select(Quotation).options(
        selectinload(Quotation.items),
    ).where(Quotation.id == quotation_id))
    return {"quotation": qt.to_dict()}


@router.delete("/quotations/{quotation_id}")
def delete_quotation(quotation_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    qt = get_or_404(db, Quotation, quotation_id, "Quotation not found")
    check_ownership(qt, user, strict=True)
    db.delete(qt)
    db.commit()
    return {"ok": True}


# --- Quotation Items ---

@router.get("/quotations/{quotation_id}/items")
def list_quotation_items(quotation_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    qt = get_or_404(db, Quotation, quotation_id, "Quotation not found")
    check_ownership(qt, user, strict=True)
    items = db.query(QuotationItem).filter_by(quotation_id=quotation_id)\
        .order_by(QuotationItem.sort_order).all()
    return {"items": [i.to_dict() for i in items]}


@router.post("/quotations/{quotation_id}/items", status_code=201)
def add_quotation_item(quotation_id: int, data: QuotationItemCreate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    qt = get_or_404(db, Quotation, quotation_id, "Quotation not found")
    check_ownership(qt, user, strict=True)
    prod = db.get(Product, data.product_id)
    snapshot = prod.to_dict() if prod else {}

    qi = QuotationItem(
        quotation_id=quotation_id,
        product_id=data.product_id,
        product_snapshot=snapshot,
        quantity=data.quantity,
        unit_price=data.unit_price if data.unit_price is not None else (float(prod.base_price or 0) if prod else 0),
        amount=data.amount,
        discount_rate=data.discount_rate,
        remark=data.remark,
        sort_order=data.sort_order,
    )
    db.add(qi)
    db.commit()
    _recalc_total(qt, db)
    db.commit()
    db.refresh(qi)
    return {"item": qi.to_dict()}


@router.put("/quotations/{quotation_id}/items/{item_id}")
def update_quotation_item(quotation_id: int, item_id: int, data: QuotationItemUpdate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    qt = get_or_404(db, Quotation, quotation_id, "Quotation not found")
    check_ownership(qt, user, strict=True)
    qi = db.get(QuotationItem, item_id)
    if not qi or qi.quotation_id != quotation_id:
        raise HTTPException(404, "Item not found")
    apply_partial_update(qi, data, ["product_id", "quantity", "unit_price", "amount", "discount_rate", "remark", "sort_order"])
    db.commit()
    qt = db.get(Quotation, quotation_id)
    _recalc_total(qt, db)
    db.commit()
    db.refresh(qi)
    return {"item": qi.to_dict()}


@router.delete("/quotations/{quotation_id}/items/{item_id}")
def delete_quotation_item(quotation_id: int, item_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    qt = get_or_404(db, Quotation, quotation_id, "Quotation not found")
    check_ownership(qt, user, strict=True)
    qi = db.get(QuotationItem, item_id)
    if not qi or qi.quotation_id != quotation_id:
        raise HTTPException(404, "Item not found")
    db.delete(qi)
    db.commit()
    _recalc_total(qt, db)
    db.commit()
    return {"ok": True}


# --- Export ---

@router.get("/quotations/{quotation_id}/export-xlsx")
def export_quotation_xlsx(quotation_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    qt = get_or_404(db, Quotation, quotation_id, "Quotation not found")
    check_ownership(qt, user, strict=True)

    import openpyxl
    from app.utils.excel_style import (
        apply_info_row, apply_title_row, apply_header_row, apply_data_row,
        apply_total_row, apply_note_row, apply_footer_row, apply_column_widths,
        num_to_chinese_uppercase, NUM_FMT_CURRENCY, NUM_FMT_NUMBER, NUM_FMT_PERCENT,
        embed_image,
    )
    from datetime import date

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报价单"
    apply_column_widths(ws)

    # Row 1: info row
    today = date.today().isoformat()
    info_text = f"公司：  |  客户：{qt.client_name or ''}  |  日期：{today}"
    apply_info_row(ws, 1, info_text)

    # Row 2: title
    title = qt.title or qt.quote_number or "报价单"
    apply_title_row(ws, 2, title)

    # Row 3: headers
    headers = ["序号", "名称", "规格型号", "型号", "功能描述", "单价", "数量", "合计", "折扣率", "成交价", "备注", "图片", "成本"]
    apply_header_row(ws, 3, headers)

    # Data rows
    items = db.query(QuotationItem).filter_by(quotation_id=quotation_id)\
        .order_by(QuotationItem.sort_order).all()
    for idx, item in enumerate(items, 1):
        snap = item.product_snapshot or {}
        qty = float(item.quantity or 0)
        price = float(item.unit_price or 0)
        discount = float(item.discount_rate or 100)
        row = 3 + idx
        formats = {6: NUM_FMT_CURRENCY, 7: NUM_FMT_NUMBER, 8: NUM_FMT_CURRENCY,
                   9: NUM_FMT_PERCENT, 10: NUM_FMT_CURRENCY}
        apply_data_row(ws, row, [
            idx,
            snap.get("name", ""),
            snap.get("model", "") or snap.get("sku", ""),
            snap.get("model", ""),
            format_description_with_specs((snap.get("description", "") or ""), snap.get("specs", {})),
            price,
            qty,
            price * qty,  # H placeholder, replaced by formula below
            discount / 100,  # I: 折扣率(小数)
            price * qty * (discount / 100),  # J placeholder
            item.remark or "",
            "",
            float(snap.get("cost_price", 0) or 0),
        ], formats)
        # Replace H and J with formulas
        ws.cell(row=row, column=8).value = f"=F{row}*G{row}"       # H: 合计 = 单价 × 数量
        ws.cell(row=row, column=10).value = f"=H{row}*I{row}"      # J: 成交价 = 合计 × 折扣率

        # Embed product image in column L
        img_url = snap.get("image_url", "")
        # Fallback: look up product's actual image_url if snapshot lost it
        if not img_url and item.product_id:
            prod = db.get(Product, item.product_id)
            if prod and prod.image_url:
                img_url = prod.image_url
        if img_url:
            import os
            from app.services.storage import UPLOAD_DIR
            if not embed_image(ws, row, 12, img_url, str(UPLOAD_DIR)):
                import logging; logging.getLogger("uvicorn").warning(f"Failed to embed image for quotation item at row {row}: {img_url}")

    # Total row
    total_row = 3 + len(items) + 1
    total_amount = float(qt.total_amount or 0)
    apply_total_row(ws, total_row, f"合计（大写）：{num_to_chinese_uppercase(total_amount)}", col_letter="J")

    # Note row
    apply_note_row(ws, total_row + 1, f"注：本报价单有效期 {qt.valid_days or 30} 天，税率 {float(qt.tax_rate or 0)}%。")

    # Footer row
    apply_footer_row(ws, total_row + 2, f"报价单编号：{qt.quote_number or ''}  |  产品数据库")

    # Log download
    from app.models.download_log import DownloadLog
    log = DownloadLog(user_id=user.id, file_type="quotation", entity_id=quotation_id, ip_address="")
    db.add(log)
    qt.download_count = (qt.download_count or 0) + 1
    db.commit()

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"quotation_{qt.quote_number or quotation_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


# --- Quotation BOM Editor ---

@router.get("/quotations/{quotation_id}/bom")
def get_quotation_bom(quotation_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    qt = get_or_404(db, Quotation, quotation_id, "Quotation not found")
    check_ownership(qt, user, strict=True)
    items = db.query(QuotationItem).filter_by(quotation_id=quotation_id)\
        .order_by(QuotationItem.sort_order).all()
    rows = []
    for idx, item in enumerate(items):
        snap = item.product_snapshot or {}
        rows.append({
            "name": snap.get("name", ""),
            "sku": snap.get("sku", ""),
            "model": snap.get("model", ""),
            "description": (snap.get("description", "") or "")[:200],
            "qty": float(item.quantity or 0),
            "price": float(item.unit_price or 0),
            "discount": float(item.discount_rate or 100),
            "remark": item.remark or "",
            "cost": float(snap.get("cost_price", 0) or 0),
        })
    return {"rows": rows, "total": float(qt.total_amount or 0)}


@router.put("/quotations/{quotation_id}/bom")
def save_quotation_bom(quotation_id: int, data: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    qt = get_or_404(db, Quotation, quotation_id, "Quotation not found")
    check_ownership(qt, user, strict=True)
    rows = data.get("rows", [])
    # Preserve old data that BOM editor doesn't carry
    old_items = {qi.sort_order: qi for qi in
                 db.query(QuotationItem).filter_by(quotation_id=quotation_id).all()}
    try:
        db.query(QuotationItem).filter_by(quotation_id=quotation_id).delete()
        for idx, row in enumerate(rows):
            old_item = old_items.get(idx + 1)
            old_snap = (old_item.product_snapshot or {}) if old_item else {}
            snap = {"name": row.get("name", ""), "sku": row.get("sku", ""),
                    "model": row.get("model", ""), "description": row.get("description", ""),
                    "cost_price": float(row.get("cost", 0) or 0)}
            for key in ("specs", "image_url", "manufacturer_name", "category_name"):
                if old_snap.get(key):
                    snap[key] = old_snap[key]
            pid = old_item.product_id if old_item else None
            qi = QuotationItem(
                quotation_id=quotation_id,
                product_id=pid,
                product_snapshot=snap,
                quantity=float(row.get("qty", 1) or 1),
                unit_price=float(row.get("price", 0) or 0),
                amount=float(row.get("qty", 1) or 1) * float(row.get("price", 0) or 0) * (float(row.get("discount", 100) or 100) / 100),
                discount_rate=float(row.get("discount", 100) or 100),
                remark=str(row.get("remark", "") or ""),
                sort_order=idx + 1,
            )
            db.add(qi)
        db.flush()
        _recalc_total(qt, db)
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(500, "BOM 保存失败，请重试")
    return {"ok": True, "total": float(qt.total_amount or 0)}
